import os
from pathlib import Path
import shutil
import sys
from PyQt6 import QtWidgets
from PyQt6.QtWidgets import QDialog, QMessageBox
from PyQt6 import QtGui
import pandas as pd
from openpyxl import Workbook, load_workbook

import settings
import convert_all
import warning_msg
import report
# import excel_app

from mainWin import Ui_MainWindow
from dialogWin import Ui_Dialog
# To get the current dirname of the absolute path
basedir = os.path.dirname(__file__)
# sets the System.AppUserModel.ID
#  It is used by Windows for taskbar icon stacking, jump lists, and taskbar pinning.
try:
    from ctypes import windll  # Only exists on Windows.
    myappid = 'ChanBangYuan.LongForm.Sc2Master.0.1'
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except ImportError:
    pass

# Establish main window
class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setFixedSize(313, 469)
        self.m_ui = Ui_MainWindow()
        self.m_ui.setupUi(self)
        # Put Signals and Events here
        self.m_ui.btn_add_sc.clicked.connect(self.button_add_Clicked)
        self.m_ui.btn_del_sc.clicked.connect(self.button_del_Clicked)
        self.m_ui.btn_convert.clicked.connect(self.btn_convert_Clicked)
        self.m_ui.btn_report.clicked.connect(self.btn_report_Clicked)
        self.m_ui.action_New.triggered.connect(self.menu_new_Clicked)
        self.m_ui.action_Open_2.triggered.connect(self.menu_open_Clicked)
        self.m_ui.action_Save_As.triggered.connect(self.menu_save_as_Clicked)
        self.m_ui.actionClose.triggered.connect(self.menu_close_Clicked)
        self.m_ui.actionChange_Default_Directory.triggered.connect(self.menu_change_dir_Clicked)
        
        self.default_path = settings.get_setting('Default Directory')
         # Check if the directory has the default template file
        new_template = self.default_path + '/template.xlsx'
        if not os.path.exists(new_template):
            # If not, copy the template file to new folderpath
            # And check if the default temolate file in the current dirtory
            old_template = os.getcwd() + '\\template.xlsx'
            if os.path.exists(old_template):
                shutil.copy(old_template, new_template)
            else:
                # notice the base template file not in current directory
                icon = QMessageBox.Icon.Warning
                title = 'Cannot find Template File'
                text = 'The base template file not found.'
                text2 = 'Find and copy the base template.xlsx into this directory'
                btns = QMessageBox.StandardButton.Ok
                ret = warning_msg.show_msg(icon, title, text, text2, btns)
        self.template_file = new_template
    
    # Excel file and app
    default_path = ''
    template_file = "template.xlsx"
    master_file = "Master.xlsx"
    app_pid = 0
    # Place global variables here
    SxS_list = []
    eps_numbers = []
    # Prepare data frames
    df_set_template = pd.DataFrame()
    df_cast_template = pd.DataFrame()
    df_cast_report = pd.DataFrame()
    df_ep = pd.DataFrame()
    time_arr = []
    # Prepare workbooks
    wb_master = Workbook()
    wb_template = Workbook()
    # For cal culate %
    total_eps = 0
    total_sc = 0
    doing_eps = 0
    doing_sc = 0
    # For check dialog box
    check_new = 1
    check_result = ""
    dlg_abort = False

    # Create the Scene class
    class Scene:
        eps = ""
        number = ""
        set = ""
        set_type = ""
        set_area = ""
        time_of_sc = ""
        time_req = 0.0
        cast_in_sc = []
        cast_in_sc_i = []
        cast_vo = []
        descriptions = ""
        extra_str = ""
        
    sc = Scene()
    # Place to hold Slots and event handling here
    # Handle file menu items
    def menu_new_Clicked(self):
        # clear or reset relevant variables
        self.template_file = "template.xlsx"
        self.master_file = "Master.xlsx"
        # Prepare data frames
        self.df_cast_report = pd.DataFrame()
        self.df_ep = pd.DataFrame()
        self.time_arr = []
        # We retain df_set_template and df_cast_template if opened
        # Clear master workbook in memory but keep template workbook
        self.wb_master = Workbook()
        # For cal culate %
        self.total_eps = 0
        self.total_sc = 0
        self.doing_eps = 0
        self.doing_sc = 0
        # For check dialog box
        self.check_new = 1
        self.check_result = ""
        self.VOchecked = ''
        self.dlg_abort = False
        self.sc = self.Scene()
        # Update Main Window display
        self.m_ui.listWidget.clear()
        self.m_ui.ledit_Master.setText(self.master_file)
        # Buttons
        self.m_ui.btn_convert.setEnabled(False)
        self.m_ui.btn_report.setEnabled(False)
        # clear status bar
        self.m_ui.statusbar.showMessage('')
        self.m_ui.statusbar.repaint()
        # clear progress bar
        self.m_ui.progressBar.setProperty("value", 0)


    def menu_open_Clicked(self):
        f_dia_return = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', 
         '/',"Excel files (*.xlsx)")
        self.master_file = f_dia_return[0]
        if self.master_file.find('master.xlsx') < 0:
            icon = QMessageBox.Icon.Warning
            title = 'Not a Master file'
            text = "This does not seem like a master file.:"
            text2 = self.master_file
            btns = QMessageBox.StandardButton.Ok
            ret = warning_msg.show_msg(icon, title, text, text2, btns)
            if ret == QMessageBox.StandardButton.Ok:
                return
        m_name = os.path.basename(self.master_file)
        self.m_ui.ledit_Master.setText(m_name)
        self.m_ui.btn_report.setEnabled(True)
        # Update status bar
        self.m_ui.statusbar.showMessage('New master set')
        self.m_ui.statusbar.repaint()
        # clear progress bar
        self.m_ui.progressBar.setProperty("value", 0)
    
    def menu_save_as_Clicked(self):
        f_dia_return = QtWidgets.QFileDialog.getSaveFileName(self, 'Save file as', 
         self.master_file,"Excel files (*.xlsx)")
        if Path(self.master_file).is_file():
            shutil.move(self.master_file, f_dia_return[0])
        else:
            icon = QMessageBox.Icon.Warning
            title = 'Abort'
            text = "Cannot find the master file at:"
            text2 = self.master_file
            btns = QMessageBox.StandardButton.Ok
            ret = warning_msg.show_msg(icon, title, text, text2, btns)
            if ret == QMessageBox.StandardButton.Ok:
                return
        self.master_file = f_dia_return[0]
        m_name = os.path.basename(self.master_file)
        self.m_ui.ledit_Master.setText(m_name)
        
    def menu_close_Clicked(self):
        self.close()
    
    def menu_change_dir_Clicked(self):
        self.default_path = settings.change_setting('Default Directory')
        
    def button_add_Clicked(self):
        f_dia_return = QtWidgets.QFileDialog.getOpenFileNames(self,
                                     "Select one or more files to open",
                                     "",
                                     "Word Document (*.docx)")
        # Extract the file list from the returned tuple, ADD to SxS_list
        self.SxS_list.extend(list(f_dia_return[0]))
        # The file list contain full path, but just want to displace file names
        display_list =  []
        for every_file in self.SxS_list:
            display_list.append(os.path.basename(every_file))
        self.m_ui.listWidget.clear()
        self.m_ui.listWidget.addItems(display_list)
        self.m_ui.btn_convert.setEnabled(True)

    def button_del_Clicked(self):
        # get the delete array
        x_found = [x.row() for x in self.m_ui.listWidget.selectedIndexes()]
        # The found list will go in the order of selection, 
        # we want to sort that by reverse ordr so that we can 
        # delect the last one first without afecting the order
        x_found.sort(reverse=True)
        # delete both the SxS_list and the list box item
        for index in x_found:
            # delete item in SxS_list
            del self.SxS_list[index]
            # Delete item on list box
            deleted = self.m_ui.listWidget.takeItem(index)
        
        if self.m_ui.listWidget.count() == 0:
            self.m_ui.btn_convert.setEnabled(False)
        else:
            self.m_ui.btn_convert.setEnabled(True)
    
    def btn_convert_Clicked(self):
        # reset all % var
        self.total_eps = 0
        self.total_sc = 0
        self.doing_eps = 0
        self.doing_sc = 0
        self.m_ui.statusbar.showMessage("Please wait, opening files")
        self.m_ui.statusbar.repaint()
        convert_all.convert_start(self)
        self.m_ui.btn_report.setEnabled(True)

    def btn_report_Clicked(self):
        # Note wb_master and wb_temp is different
        # check if master file exit
        wbs = report.est_wb(self)
        if wbs == 'Aborted':
            return
        wb_temp = wbs[0]
        wb_master = wbs[1]
        df_team = wbs[2]
        self.m_ui.statusbar.showMessage("Reporting started.")
        self.m_ui.statusbar.repaint()
        all_eps = report.group_to_team(wb_master, df_team)
        if all_eps == 'Aborted':
            return
        self.m_ui.statusbar.showMessage("Consolidating masters into teams")
        self.m_ui.statusbar.repaint()
        report.prepare_team_master(self, wb_master, all_eps)
        self.m_ui.statusbar.showMessage("Reporting casts.")
        self.m_ui.statusbar.repaint()
        self.m_ui.progressBar.setValue(80)
        report.report_cast(self, wb_master, all_eps)
        self.m_ui.statusbar.showMessage("Reporting sets.")
        self.m_ui.statusbar.repaint()
        self.m_ui.progressBar.setValue(90)
        report.report_sets(self, wb_master, all_eps)
        self.m_ui.progressBar.setValue(100)
        self.m_ui.statusbar.showMessage("Done reporting")
        self.m_ui.statusbar.repaint()
        # wb_master.save()
        # clean_up.clean_excel(self)
        

    # The function to launch the check dialog box
    # Usage: result = launch_ck_dialog_box(title, note, to_find, ck_list)
    # Parameters: 
    # title =   dialog box's title
    # note =    instruction note
    # to_find = subject to find in:
    # ck_list = the list of options
    # result =  New?(boolean), result string
    def launch_ck_dialog_box(self, title, where, note, to_find, ck_list):
        dlg = Check_dialog(parent=self, title=title, where=where, note=note, ck_list=ck_list, to_find=to_find)
        dlg.exec()
        return dlg.ck_result()

    #def closeEvent(self, event):
    #    clean_up.clean_excel(self)
        
    
class Check_dialog(QDialog):
     # set global variable
    ep_sc = ''
    # the check dialog box
    def __init__(self, parent=None, title=None, where=None, note=None, ck_list=None, to_find=None):
        super().__init__(parent)
        # Create an instance of the GUI
        self.d_ui = Ui_Dialog()
        # Run the .setupUi() method to show the GUI
        self.d_ui.setupUi(self)
        
        # set new position
        self.setGeometry(parent.geometry().x() + self.width() + 13,parent.geometry().y(),self.width(),self.height())
        self.setFixedSize(293, 421)
        self.check_new = 0
        self.VOchecked = ''
        self.check_result = to_find
        self.ep_sc = where[where.find('#'):where.find(' ', where.find('#'))]
        self.setWindowTitle(title)
        if title == "Found new set":
            self.d_ui.label_Title.setText('Set')
            self.d_ui.Button_Extra.hide()
            self.d_ui.VOcheckBox.hide()
        elif title == "Found new area":
            self.d_ui.label_Title.setText('Area')
            self.d_ui.Button_Extra.hide()
            self.d_ui.VOcheckBox.hide()
        elif title == "Found new cast":
            self.d_ui.label_Title.setText('Cast')
            self.d_ui.Button_Extra.show()
            self.d_ui.VOcheckBox.show()

        self.d_ui.lineEdit.setText(to_find)
        self.d_ui.label_2.setText(where)
        self.d_ui.label.setText(note)
        self.d_ui.listWidget.addItems(ck_list)
        self.d_ui.VOcheckBox.isChecked = False
        
        self.d_ui.Button_deselect.clicked.connect(self.dlg_deslect_Clicked)
        self.d_ui.Button_New.clicked.connect(self.dlg_new_Clicked)
        self.d_ui.listWidget.doubleClicked.connect(self.dlg_listed_Clicked)
        self.d_ui.Button_Listed.clicked.connect(self.dlg_listed_Clicked)
        self.d_ui.listWidget.itemSelectionChanged.connect(self.list_current_item_changed)
        self.d_ui.Button_Listed.setEnabled(False)
        self.d_ui.Button_New.setEnabled(True)
        self.d_ui.Button_Extra.clicked.connect(self.dlg_extra_Clicked)
        self.d_ui.Button_Abort.clicked.connect(self.dlg_abort_Clicked)
        self.d_ui.VOcheckBox.toggled.connect(self.item_selected)

    def dlg_deslect_Clicked(self):
        items_selected = len(self.d_ui.listWidget.selectedItems())
        if items_selected > 0 :
            self.d_ui.listWidget.currentItem().setSelected(False)
    def closeEvent(self, event):
        # accept() close the dialog while ignor() will do nothing, not even getting out.
        event.accept()
    def list_current_item_changed(self):
        items_selected = len(self.d_ui.listWidget.selectedItems())
        if items_selected > 0 :
            self.d_ui.Button_Listed.setEnabled(True)
            self.d_ui.Button_New.setEnabled(False)
        else:
            self.d_ui.Button_Listed.setEnabled(False)
            self.d_ui.Button_New.setEnabled(True)

    def dlg_new_Clicked(self):
        self.check_new = 1
        self.check_result = self.d_ui.lineEdit.text()
        self.done(1)

    def dlg_listed_Clicked(self):
        self.check_new = 0
        self.check_result = self.d_ui.listWidget.currentItem().text()
        self.done(1)

    def dlg_extra_Clicked(self):
        self.check_new = 2
        self.check_result = self.d_ui.lineEdit.text()
        self.done(1)

    def dlg_abort_Clicked(self):
        icon = QMessageBox.Icon.Critical
        title = 'Abort'
        text = 'Abort the conversion?'
        text2 = 'Converted up to ' + self.ep_sc
        btns = QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel
        ret = warning_msg.show_msg(icon, title, text, text2, btns)
        if ret == QMessageBox.StandardButton.Ok:
            self.check_new = 3
        elif ret == QMessageBox.StandardButton.Cancel:
            return
        self.done(1)

    def item_selected(self):
        self.VOchecked = 'VO'
       
    def ck_result(self):
        return self.check_new, self.check_result, self.VOchecked

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon(os.path.join(basedir, 's2m.ico')))
    window = MainWindow()
    window.show()
    app.exec()
