#! python3
from PyQt6.QtWidgets import QMessageBox
from docx2python import docx2python
from pathlib import Path
import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import convert_ep
import format
import warning_msg

def convert_start(self):
    self.m_ui.statusbar.showMessage("Conversion started")
    self.m_ui.statusbar.repaint()
    # Check all files in SxS_list
    # since all files are from the add button, presuming they are OK
    self.eps_numbers = []
    for each_ep in self.SxS_list:
        try:
            s = docx2python(each_ep).text
        except Exception as e:
            icon = QMessageBox.Icon.Warning
            title = 'Cannot open file:'
            text = e.strerror + ', Check if file opened'
            text2 = os.path.basename(e.filename)
            btns = QMessageBox.StandardButton.Ok
            ret = warning_msg.show_msg(icon, title, text, text2, btns)
            return 'CovertedNone'
        # Find episode number
        search_start = re.search("Episode", s, re.IGNORECASE).end()
        search_stop = s.find("\n", search_start)
        this_ep_number = s[search_start:search_stop].strip()
        self.eps_numbers.append(this_ep_number)
    # For calculate %
    self.total_eps = len(self.eps_numbers)
    # Generate master file name only if it's blank or new file
    if self.m_ui.ledit_Master.text() == "" or self.m_ui.ledit_Master.text() == "New Master.xlsx":
        if len(self.eps_numbers) == 1:
            self.master_file = self.default_path + '/' + self.eps_numbers[0] + " master.xlsx"
            self.m_ui.ledit_Master.setText(os.path.basename(self.master_file))
        else:
            self.master_file = self.default_path + '/' + self.eps_numbers[0] + " to " + self.eps_numbers[len(self.eps_numbers)-1] + " master.xlsx"
            self.m_ui.ledit_Master.setText(os.path.basename(self.master_file))
       
    # Setting up all data frames
    # At the launch of the program, 
    # template file exist is checked and never changed, 
    # no need to check again.
    # All df has been established in main, only load if empty
    self.m_ui.statusbar.showMessage("Loading data from template file")
    self.m_ui.statusbar.repaint()
    if self.df_ep.shape[1] == 0:
        # df_eps has 0 columns, open from template
        self.df_ep = pd.read_excel(self.template_file, sheet_name = 'Eps')
    elif self.df_ep.shape[0] != 0:
        # df_ep has scenes entry, clear
        self.df_ep = pd.DataFrame(columns=self.df_ep.columns)

    if self.df_set_template.shape[1] == 0:
        # df_set_template has 0 columns, open from template
        self.df_set_template = pd.read_excel(self.template_file, sheet_name = 'Sets')

    if self.df_cast_template.shape[1] == 0:
        # df_cast_template has 0 columns, open from template
        self.df_cast_template  = pd.read_excel(self.template_file, sheet_name = 'Casts')
        # Change all 'New' to 'RC'
        self.df_cast_template.loc[self.df_cast_template['Type'] == 'New', 'Type'] = 'RC'
        # Drop duplicated rows if any
        self.df_cast_template = self.df_cast_template.drop_duplicates(subset=['Cast'])

    if len(self.time_arr) == 0:
        # establish time array 
        df_time = pd.read_excel(self.template_file, sheet_name = 'Time')
        self.time_arr = df_time['Time'].tolist()
    
    # Check if need to load Excel files 
    if len(self.wb_template.sheetnames) == 1 and self.wb_template.sheetnames == ['Sheet']:
        # template.xlsx not loaded, to load
        self.wb_template = load_workbook(self.template_file)

    # check if the master file exist
    path = Path(self.master_file)
    if path.is_file():
        # Open the file
        self.m_ui.statusbar.showMessage("Loading master file")
        self.m_ui.statusbar.repaint()
        self.wb_master = load_workbook(self.master_file)
    else:
        # if no existing master file, create and save one.
        self.m_ui.statusbar.showMessage("Creating and saving master file")
        self.m_ui.statusbar.repaint()
        self.wb_master = Workbook()
        self.wb_master.save(self.master_file)
    if 'Cast Report' in self.wb_master.sheetnames:
        self.df_cast_report = pd.read_excel(self.master_file, sheet_name = 'Cast Report')
        if self.df_cast_report.shape[1] > 2:
            # Has done report before, discard other columns
            self.df_cast_report.drop(self.df_cast_report.columns[2:self.df_cast_report.shape[1]], axis = 1, inplace = True)
            # delete the multi heading
            self.df_cast_report.drop(self.df_cast_report.index[0], axis= 0, inplace = True)
            # delete all sheets except masters and cast report
            for sheet in self.wb_master.sheetnames:
                if not sheet.isdigit() and sheet != 'Cast Report':
                    del self.wb_master[sheet]
    else:
        # create a new one
        self.df_cast_report = pd.DataFrame(columns = self.df_cast_template.columns)
        # fill the main cast
        self.df_cast_report = self.df_cast_template[self.df_cast_template.Type == 'Main'].copy()
        # drop the Artises column
        self.df_cast_report.drop(self.df_cast_report.columns[2], axis = 1, inplace = True)
    # Going in each episode
    for i in range(len(self.SxS_list)):
        if self.dlg_abort:
            self.dlg_abort = False
            break
        try:
            s = docx2python(self.SxS_list[i]).text
        except Exception as e:
            icon = QMessageBox.Icon.Warning
            title = 'Cannot open file:'
            text = e.strerror + ', Check if file opened'
            text2 = os.path.basename(e.filename)
            btns = QMessageBox.StandardButton.Ok
            ret = warning_msg.show_msg(icon, title, text, text2, btns)
            break
        # check if footnote or endnote at the end, 
        # this will cause error in finding next scene
        # because they appear as footnote1)/t and endnote1)/t
        # the extration will give both as scene e1
        if s.find('footnote') != -1:
            s = s[0:s.find('footnote')]
        if s.find('endnote') != -1:
            s = s[0:s.find('endnote')]
        self.sc.eps = self.eps_numbers[i]
        self.doing_eps = i + 1
        self.m_ui.statusbar.showMessage("Converting #" + self.sc.eps)
        self.m_ui.statusbar.repaint()
        ret = convert_ep.convert_this_ep(self, s)
        if ret == 'abort':
            # return to caller function master
            return
        
        # Update the master dataframe column heading with new casts
        cast_arr = list(self.df_cast_report['Cast'])
        # it is a single heading for cast report before reporting
        # but after reporting, this df is double heading, so cut first null entry
        if self.df_cast_report['Cast'].isnull().values.any():
            cast_arr = cast_arr[1:len(cast_arr)]
    
        col_arr = list(self.df_ep)
        for name in cast_arr:
            col_arr[10 + cast_arr.index(name)] = name
        self.df_ep.columns = col_arr

        # Update Excel master file here
        # Delete old and write new sheet to avoid creating extra sheet with name + '1'
        ws_array = self.wb_master.sheetnames
        if self.sc.eps in ws_array:
            # mark the position of the sheet in wb
            sheet_index = ws_array.index(self.sc.eps)
            del self.wb_master[self.sc.eps]
        else:
            # Add in the sheet list to find position of the sheet
            ws_array.append(self.sc.eps)
            ws_array.sort()
            sheet_index = ws_array.index(self.sc.eps)
            
        ws_thisEp = self.wb_master.create_sheet(self.sc.eps, sheet_index)
        # check and delete the default sheet "Sheet"
        if len(ws_array) > 1 and 'Sheet' in ws_array:
            del self.wb_master['Sheet']
            ws_array.remove('Sheet')
            ws_array.sort()
            sheet_index = ws_array.index(self.sc.eps)
        
        # copy data over
        for r in dataframe_to_rows(self.df_ep, index=False, header=True):
            ws_thisEp.append(r)

        # Get the number of cast used from df_cast_report
        CastUsed = self.df_cast_report.shape[0]
        if self.df_cast_report['Cast'].isnull().values.any():
            # this is from the reported cast report, multi headings, -1 to adjust
            CastUsed = CastUsed - 1
        # format the master
        ws_thisEp = format.master_sheet(ws_thisEp, CastUsed)

        # update cast ws in master
        if 'Cast Report' in ws_array:
            # mark the position of the sheet in wb
            sheet_index = ws_array.index('Cast Report')
            del self.wb_master['Cast Report']
        else:
            # Add in the sheet list to find position of the sheet
            ws_array.append('Cast Report')
            ws_array.sort()
            sheet_index = ws_array.index('Cast Report')
            
        ws_thisCast_r = self.wb_master.create_sheet('Cast Report', sheet_index)
        # this always create a clean report as new
        for r in dataframe_to_rows(self.df_cast_report, index=False, header=True):
            ws_thisCast_r.append(r)
        # Format 
        ws_thisCast_r =format.cast(ws_thisCast_r)

        # Update cast and set in wb template
        ws_array_temp = self.wb_template.sheetnames
        if 'Casts' in ws_array_temp:
            del self.wb_template['Casts']

        ws_thisCast_t = self.wb_template.create_sheet('Casts', 1)
        for r in dataframe_to_rows(self.df_cast_template, index=False, header=True):
            ws_thisCast_t.append(r)

        # 'Casts' need format
        ws_thisCast_t = format.cast(ws_thisCast_t)

        if 'Sets' in ws_array_temp:
            del self.wb_template['Sets']

        ws_thisSet_t = self.wb_template.create_sheet('Sets', 2)
        for r in dataframe_to_rows(self.df_set_template, index=False, header=True):
            ws_thisSet_t.append(r)
        
        # 'Sets' need format
        ws_thisSet_t = format.set(ws_thisSet_t)

        # at the end of each episode, save both excel files
        self.m_ui.statusbar.showMessage("Updating #" + self.sc.eps)
        self.m_ui.statusbar.repaint()
        self.wb_master.save(self.master_file)
        self.wb_template.save(self.template_file)
        # Clean up df_ep for next episode
        if self.df_ep.shape[0] != 0:
            self.df_ep = pd.DataFrame(columns = self.df_ep.columns)
       
    self.m_ui.statusbar.showMessage("All episodes converted.")
    self.m_ui.statusbar.repaint()
    
   
