# from curses import start_color
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def master_sheet(ws, CastUsed):
    # Formatting master only
    # presuming pd has loaded into the ws
    # return ws without saving
    lastrow = ws.max_row
    lastcolumn = ws.max_column

    # Border
    for col in ws.iter_cols(min_row=1, max_col= lastcolumn, max_row=lastrow):
        for cell in col:
            cell.border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # Row height
    ws.row_dimensions[1].height = 88
    for row in range(2, lastrow + 1):
        ws.row_dimensions[row].height = 60

    col = 1     #Time
    ws.column_dimensions[get_column_letter(col)].width = 14.09
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='center', 
                wrap_text=True)
            
    col = 2     #Ep
    ws.column_dimensions[get_column_letter(col)].width = 4.91
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
            
    col = 3     #Sc
    ws.column_dimensions[get_column_letter(col)].width = 4.64
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
            
    col = 4     #Set
    ws.column_dimensions[get_column_letter(col)].width = 28.45
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='left', vertical='top', 
                wrap_text=True)
            
    col = 5     #Area
    ws.column_dimensions[get_column_letter(col)].width = 31.55
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='left', vertical='top', 
                wrap_text=True)
            
    col = 6     #D/N
    ws.column_dimensions[get_column_letter(col)].width = 17.09
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
            
    col = 7     #Type
    ws.column_dimensions[get_column_letter(col)].width = 7.82
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
            
    col = 8     #Shoot Time
    ws.column_dimensions[get_column_letter(col)].width = 8.18
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.number_format = 'h:mm'
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
            
    col = 9     #Pages
    ws.column_dimensions[get_column_letter(col)].width = 7.73
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
            
    col = 10     #Synopsis
    ws.column_dimensions[get_column_letter(col)].width = 47
    ws.cell(row = 1, column = col).font = Font(size=14)
    ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
        vertical='top', wrap_text=True)
    for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
        for cell in col:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='left', vertical='top', 
                wrap_text=True)

    # Now for the cast columns
    for col in range(11, 73):
        ws.column_dimensions[get_column_letter(col)].width = 3.09
        
    for row in ws.iter_rows(min_col=11, max_col=72, min_row=1, max_row= 1):
        for cell in row:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                textRotation=180, wrap_text=True)
    for row in ws.iter_rows(min_col=11, max_col=72, min_row=2, max_row=lastrow):
        for cell in row:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)

    # Last 3 columns
    for col in range(73, 75 + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26.64
    for row in ws.iter_rows(min_col=73, max_col=75, min_row=1, max_row= 1):
        for cell in row:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='center', vertical='top', 
                wrap_text=True)
    for row in ws.iter_rows(min_col=73, max_col=75, min_row=2, max_row=lastrow):
        for cell in row:
            cell.font = Font(size=14)
            cell.alignment=Alignment(horizontal='left', vertical='top', 
                wrap_text=True)
    # Hide empty column for master only
    if ws.title.isdigit():
        for col in range(11, CastUsed + 11):
            column_empty = True
            for row_num in range(2, lastrow + 2):
                if ws.cell(row=row_num, column=col).value == 'X' or ws.cell(row=row_num, column=col).value == 'Y':
                    #print(str(col) + ' NotXX ' + ws.cell(row=1, column=col).value)
                    column_empty = False
                    break
            if column_empty:
                #print(str(col) + ' Empty ' + ws.cell(row=1, column=col).value)
                ws.column_dimensions[get_column_letter(col)].hidden= True
    # Hide unused part timers column
    for col in range(11 + CastUsed, 73):
        ws.column_dimensions[get_column_letter(col)].hidden= True

    # zoom to 50%
    ws.sheet_view.zoomScale = 70 
    return ws
    

def set(ws):
    # Formatting master only
    # presuming pd has loaded into the ws
    # return ws without saving
    lastrow = ws.max_row
    lastcolumn = ws.max_column

    for row in ws.iter_rows(min_col=1, max_col=lastcolumn, min_row=1, max_row=lastrow):
        for cell in row:
            cell.font = Font(size=12)
            cell.alignment=Alignment(horizontal='left', vertical='top')

    ws.column_dimensions[get_column_letter(1)].width = 5.17
    ws.column_dimensions[get_column_letter(2)].width = 33.67
    for col in range(3, lastcolumn + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24.17
    return ws

def cast(ws):
    # Formatting master only
    # presuming pd has loaded into the ws
    # return ws without saving
    lastrow = ws.max_row
    lastcolumn = ws.max_column

    for row in ws.iter_rows(min_col=1, max_col=lastcolumn, min_row=1, max_row=lastrow):
        for cell in row:
            cell.font = Font(size=12)
            cell.alignment=Alignment(horizontal='left', vertical='top')

    ws.column_dimensions[get_column_letter(1)].width = 6.67
    ws.column_dimensions[get_column_letter(2)].width = 11.33
    ws.column_dimensions[get_column_letter(3)].width = 14.25
    return ws

def cast_final(ws, team_list):
    # Formatting cast report after reporting
    # presuming pd has loaded into the ws
    # return ws without saving
    lastrow = ws.max_row
    lastcolumn = ws.max_column
    
    # Rewrite first column since dataframe to columns handle multi index differently
    ws.cell(row = 1, column = 2).value = 'Cast'
    col = 3
    for team_name in team_list:
        ws.cell(row = 1, column = col).value = team_name
        ws.cell(row = 1, column = col + 1).value = ''
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        col = col + 2
    ws.cell(row = 1, column = col).value = 'Total'
    ws.cell(row = 1, column = col + 1).value = ''
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
    # col width
    ws.column_dimensions[get_column_letter(1)].width = 6.67
    ws.column_dimensions[get_column_letter(2)].width = 13.45
    for col in range(3, lastcolumn + 1):
        ws.column_dimensions[get_column_letter(col)].width = 5.36
    # bold
    for row in ws.iter_rows(min_col=lastcolumn - 1, max_col=lastcolumn + 1, min_row=1, max_row=lastrow):
        for cell in row:
            cell.font = Font(bold=True)
    # center for all row except first, use default as 'gereral'
    for row in ws.iter_rows(min_col=3, max_col=lastcolumn + 1, min_row=1, max_row=lastrow):
        for cell in row:
            cell.alignment=Alignment(horizontal='center')
    # borders
    # generate a list of row need border
    border_row = list(range(2, lastrow + 1, 5))
    for b_row in border_row:
        for col in ws.iter_cols(min_row=b_row, max_col= lastcolumn, max_row=b_row):
            for cell in col:
                cell.border = Border(bottom=Side(style='thin'))

    # grey all '#Set'
    grey_col = list(range(4, lastcolumn + 1, 2))
    for g_col in grey_col:
        for r in ws.iter_rows(min_row=3, max_row = lastrow, min_col = g_col, max_col=g_col):
            for cell in r:
                cell.fill = PatternFill(start_color='00E6E6E6', end_color='00E6E6E6', fill_type='solid')

def set_final(ws):
    # Formatting master only
    # presuming pd has loaded into the ws
    lastrow = ws.max_row
    lastcolumn = ws.max_column
    last_set = lastrow - 7

    # col width
    ws.column_dimensions[get_column_letter(1)].width = 4.45
    ws.column_dimensions[get_column_letter(2)].width = 27.36
    for col in range(3, lastcolumn + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7
    # center for all row except first 2 col, use default as 'gereral'
    for row in ws.iter_rows(min_col=3, max_col=lastcolumn + 1, min_row=1, max_row=last_set):
        for cell in row:
            cell.alignment=Alignment(horizontal='center')
     # bold
    for row in ws.iter_rows(min_col=lastcolumn, max_col=lastcolumn, min_row=1, max_row=last_set):
        for cell in row:
            cell.font = Font(bold=True)
            # borders
    # generate a list of row need border
    border_row = list(range(1, last_set + 1, 5))
    for b_row in border_row:
        for col in ws.iter_cols(min_row=b_row, max_col= lastcolumn, max_row=b_row):
            for cell in col:
                cell.border = Border(bottom=Side(style='thin'))
    # Total with double border
    for col in ws.iter_cols(min_row=last_set + 1, max_row=last_set + 1, min_col = 1, max_col= lastcolumn):
        for cell in col:
            cell.alignment=Alignment(horizontal='center')
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))
    # format percentage
    for r in ws.iter_rows(min_row=lastrow - 2, max_row = lastrow, min_col = 4, max_col=4):
        for cell in r:
            cell.number_format = '0.00%'

def schedule(ws):
    # Formatting master only
    # presuming pd has loaded into the ws
    lastrow = ws.max_row
    lastcolumn = ws.max_column
    # heading row
    for row in ws.iter_rows(min_col=1, max_col=lastcolumn, min_row=1, max_row= 1):
        for cell in row:
            cell.alignment=Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

    # col 1 Day
    ws.column_dimensions[get_column_letter(1)].width = 3.82
    for col in ws.iter_cols(min_col= 1, max_col= 1, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(horizontal='center', vertical='top')
    # col 2 Date
    ws.column_dimensions[get_column_letter(2)].width = 16.09
    for col in ws.iter_cols(min_col= 2, max_col= 2, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(horizontal='right', vertical='top')
            cell.number_format = 'dd-mmm-yy ddd'
    # col 3 and 4, Team and Loc
    ws.column_dimensions[get_column_letter(3)].width = 5
    ws.column_dimensions[get_column_letter(4)].width = 5
    for col in ws.iter_cols(min_col= 3, max_col= 4, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(horizontal='center', vertical='top')
    # col 5 Set
    ws.column_dimensions[get_column_letter(5)].width = 30.09
    for col in ws.iter_cols(min_col= 5, max_col= 5, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(horizontal='left', vertical='top')
    # col 6,7 From, To timing
    ws.column_dimensions[get_column_letter(6)].width = 6.82
    ws.column_dimensions[get_column_letter(7)].width = 6.82
    for col in ws.iter_cols(min_col= 6, max_col= 7, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(vertical='top')
            cell.number_format = 'hh:mm'
    # col 8,9 Sc and Pg
    ws.column_dimensions[get_column_letter(8)].width = 5.73
    ws.column_dimensions[get_column_letter(9)].width = 5.73
    for col in ws.iter_cols(min_col= 8, max_col= 9, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(vertical='top')
    # col 10 Time
    ws.column_dimensions[get_column_letter(10)].width = 6.82
    for col in ws.iter_cols(min_col= 10, max_col= 10, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(vertical='top')
            cell.number_format = 'hh:mm'
    # col 11, 12 'Extra' and 'Meal'
    ws.column_dimensions[get_column_letter(11)].width = 5.55
    ws.column_dimensions[get_column_letter(12)].width = 5.55
    for col in ws.iter_cols(min_col= 11, max_col= 12, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(vertical='top')
    # col 13 to 15 the three remarks columns
    ws.column_dimensions[get_column_letter(13)].width = 27.45
    ws.column_dimensions[get_column_letter(14)].width = 27.45
    ws.column_dimensions[get_column_letter(15)].width = 27.45
    for col in ws.iter_cols(min_col= 13, max_col= 15, min_row=2, max_row=lastrow):
        for cell in col:
            cell.alignment=Alignment(vertical='top', wrap_text=True)

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col= lastcolumn):
        for cell in col:
            cell.border = Border(bottom=Side(style='thin'), top=Side(border_style='thick'))
    for col in ws.iter_cols(min_row=lastrow, max_row=lastrow, min_col=1, max_col= lastcolumn):
        for cell in col:
            cell.border = Border(bottom=Side(style='thick'))
    