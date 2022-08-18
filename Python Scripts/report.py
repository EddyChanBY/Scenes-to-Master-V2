
from pathlib import Path
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt6.QtWidgets import QDialog, QMessageBox
import warning_msg
import format
# to do reporting

def est_wb(self):
    # Check if the master listed in the text box exist
    path = Path(self.master_file)
    if not path.is_file():
        icon = QMessageBox.Icon.Critical
        title = 'Abort'
        text = "Cannot find the template file"
        text2 = 'Please abort.'
        btns = QMessageBox.StandardButton.Ok
        ret = warning_msg.show_msg(icon, title, text, text2, btns)
        if ret == QMessageBox.StandardButton.Ok:
            return 'Aborted'
    wb_temp = load_workbook(self.template_file)
    wb_master = load_workbook(self.master_file)
    df_team = pd.read_excel(self.template_file, sheet_name='Teams')
    if self.df_set_template.shape[1] == 0:
        # df_set_template has 0 columns, open from template
        self.df_set_template = pd.read_excel(self.template_file, sheet_name = 'Sets')
    # presuming by reporting time the 'Cast Report must be in master
    self.df_cast_report = pd.read_excel(self.master_file, sheet_name = 'Cast Report')
    return[wb_temp, wb_master, df_team]

def group_to_team(wb_master, df_team):
    ws_array = wb_master.sheetnames
    team_array = []
    all_eps = {}
    # tested each_ep is str 
    for each_ep in ws_array:
        if each_ep.isdigit():
            try:
                team_name = df_team.loc[(df_team['From'] <= int(each_ep)) 
                    & (df_team['To'] >= int(each_ep)), 'Team'].values[0]
            except:
                icon = QMessageBox.Icon.Warning
                title = 'Abort'
                text = 'It seems Episode ' + each_ep + ' is not in the team schedule.'
                text2 = 'Need to abort now!'
                btns = QMessageBox.StandardButton.Ok 
                ret = warning_msg.show_msg(icon, title, text, text2, btns)
                return 'Aborted'

            if team_name not in team_array:
                all_eps[team_name] = [each_ep]
                team_array.append(team_name)
            else:
                all_eps[team_name].append(each_ep)
    return all_eps

def prepare_team_master(self, wb_master, all_eps):
    ep_done = 0
    total_eps = 0
    # Update the master dataframe column heading with new casts
    cast_arr = list(self.df_cast_report['Cast'])
    # it is a single heading for cast report for the first time
    # but for any redo for reporting, it is double heading, so cut first null entry
    if self.df_cast_report['Cast'].isnull().values.any():
        cast_arr = cast_arr[1:len(cast_arr)]
    # if there are differt names in header, a new column will be added when concat
    # prepare standard header and overwrite each ep master so they have common columns
    df_team_m = pd.read_excel(self.template_file, sheet_name='Eps')
    col_arr = list(df_team_m)
    for index, name in enumerate(cast_arr):
        col_arr[10 + index] = name
    
    for eps in all_eps:
        total_eps = total_eps + len(all_eps[eps])
    for team in all_eps:
        team_eps = all_eps[team]
        # Clear Team master
        df_team_m = pd.DataFrame()
        for ep in team_eps:
            if df_team_m.shape[0] == 0:
                # set first ep in df
                df_team_m = pd.read_excel(self.master_file, sheet_name=ep)
                df_team_m.columns = col_arr
            else:
                # df exist and just append
                df_ep = pd.read_excel(self.master_file, sheet_name=ep)
                df_ep.columns = col_arr
                df_team_m = pd.concat([df_team_m, df_ep], axis=0, ignore_index = True)

            ep_done = ep_done + 1
        # all eps for this team added to df
        df_team_m = df_team_m.sort_values(['Type', 'Set'], ascending = [False, True])
        # save df_team_m in master.xlsx with name like 'Team A'
        sheets = wb_master.sheetnames
        team_name = 'Team ' + team
        if team_name in sheets:
            del wb_master[team_name]
        ws_this_team = wb_master.create_sheet(team_name)
        # copy data over
        for r in dataframe_to_rows(df_team_m, index=False, header=True):
            ws_this_team.append(r)
        # Format
        ws_this_team = format.master_sheet(ws_this_team, len(cast_arr))
        # Save
        wb_master.save(self.master_file)
        percent_done = int(ep_done/total_eps * 80)
        self.m_ui.progressBar.setValue(percent_done)
        
def report_sets(self, wb_master, all_eps):
    # because of similarity, we do the Scheduling report as well in this module
    # prepare a 2 dimentional list with ['Type', 'Set'] for each member
    set_list = list(map(list, zip(list(self.df_set_template['Type']), list(self.df_set_template['Set']))))
    team_list = []
    col_list = ['Type', 'Set']
    for team in all_eps.keys():
        team_list.append('Team ' + team)
        col_list.append('Team ' + team)
    col_list.append('Total')
    # prepare df_report for set
    df_report = pd.DataFrame(columns = col_list)
    df_report['Type'] = self.df_set_template['Type']
    df_report['Set'] = self.df_set_template['Set']
    # prepare df_schedule
    df_schedule = pd.DataFrame(columns = ['Day', ' Date', 'Team', 'Loc', 'Set', 'From', 'To', 'Sc'])
    # check data in team master
    for index,team in enumerate(team_list):
        col = index + 2
        team_short = team[5:]
        df_team_all = pd.read_excel(self.master_file, sheet_name = team)
        for row, set in enumerate(set_list):
            appear = (df_team_all.Set == set[1]).sum()
            df_report.iat[row, col] = appear
            df_schedule.loc[len(df_schedule.index)] = ['', '', team_short, set[0], set[1], '', '', appear]
    # Calculate total
    for row, set_name in enumerate(set_list):
        total_sc = 0
        col = 2
        for index, team in enumerate(team_list):
            total_sc = total_sc + df_report.iat[row, col]
            col = col + 1
        df_report.iat[row, col] = total_sc
    # Do scene total
    row = len(df_report.index)
    col = 2
    for team in team_list:
        df_report.loc[row, team] = df_report[team].sum()
    df_report.loc[row, 'Total'] = df_report['Total'].sum()
    # Do percentage
    total_ST = df_report.loc[df_report.Type == 'ST', 'Total'].sum()
    total_RC = df_report.loc[df_report.Type == 'RC', 'Total'].sum()
    total_OB = df_report.loc[df_report.Type == 'OB', 'Total'].sum()
    total_set = total_ST + total_RC + total_OB
    # Delete all rows with 0 in total
    df_report = df_report.drop(df_report[df_report.Total == 0].index)
    df_report = df_report.reset_index(drop=True)
    # add summary
    col_pad = ['']*(df_report.shape[1] - 4)
    # Add an empty row
    df_report.loc[df_report.shape[0]] = ['']*(df_report.shape[1])
    df_report.loc[df_report.shape[0]] = ['Statistic:'] + ['']*(df_report.shape[1] - 1)
    df_report.loc[df_report.shape[0]] = ['', 'Descriptions', 'Count', 'Percentage'] + col_pad
    df_report.loc[df_report.shape[0]] = ['ST', 'Studio Sets', total_ST, (total_ST / total_set)]+ col_pad
    df_report.loc[df_report.shape[0]] = ['RC', 'Recurrent OB Sets', total_RC, (total_RC / total_set)]+ col_pad
    df_report.loc[df_report.shape[0]] = ['OB', 'Outside Broadcast', total_OB, (total_OB / total_set)]+ col_pad
    
    sheets = wb_master.sheetnames
    if 'Set Report' in sheets:
        del wb_master['Set Report']
    ws_set_report = wb_master.create_sheet('Set Report')
    for r in dataframe_to_rows(df_report, index=False, header=True):
        ws_set_report.append(r)
    # Format
    ws_set_report = format.set_final(ws_set_report)
    
    # Delete all rows with 0 in total
    df_schedule = df_schedule.drop(df_schedule[df_schedule.Sc == 0].index)
    if 'Scheduling Summary' in sheets:
        del wb_master['Scheduling Summary']
    ws_schedule_report = wb_master.create_sheet('Scheduling Summary')
    for r in dataframe_to_rows(df_schedule, index=False, header=True):
        ws_schedule_report.append(r)
    # Format then
    ws_schedule_report = format.schedule(ws_schedule_report)
    # Save
    wb_master.save(self.master_file)

    '''# Format sheet
    ws_set = wb_master.sheets['Set Report']
    set_maxr = ws_set.range('B3').end('down').row + 1
    set_maxc = ws_set.range('A1').end('right').column
    # Set the column width
    ws_set.range((1, 2),(set_maxr, 1)).autofit()
    ws_set.range((1, 3),(set_maxr,set_maxc)).column_width = 7
    # Set alignment
    ws_set.range((1,3),(set_maxr,set_maxc)).api.HorizontalAlignment = -4108
    # Bold the total values
    ws_set.range((1,set_maxc),(set_maxr,set_maxc)).api.Font.Bold = True
    # Draw border
    row = 2
    while row < set_maxr:
        ws_set.range((row,1),(row,set_maxc)).api.Borders(8).LineStyle = 1
        ws_set.range((row,1),(row,set_maxc)).api.Borders(8).Weight = 2
        row = row + 5
    ws_set.range((set_maxr,1),(set_maxr,set_maxc)).api.Borders(8).LineStyle = 1
    ws_set.range((set_maxr,1),(set_maxr,set_maxc)).api.Borders(8).Weight = 2
    ws_set.range((set_maxr,1),(set_maxr,set_maxc)).api.Borders(9).LineStyle = -4119
    ws_set.range((set_maxr,1),(set_maxr,set_maxc)).api.Borders(9).Weight = 2
    # display percentage
    row = set_maxr + 2
    total_set = total_ST + total_RC + total_OB
    ws_set.range((row,1)).value = 'Statistic:'
    ws_set.range((row + 1,2)).value = 'Descriptions'
    ws_set.range((row + 1,3)).value = 'Count'
    ws_set.range((row + 1,4)).value = 'Percentage'
    ws_set.range((row + 2,1)).value = 'ST'
    ws_set.range((row + 2,2)).value = 'Studio Sets'
    ws_set.range((row + 2,3)).value = total_ST
    ws_set.range((row + 2,4)).value = (total_ST / total_set)
    ws_set.range((row + 3,1)).value = 'RC'
    ws_set.range((row + 3,2)).value = 'Recurrent OB Sets'
    ws_set.range((row + 3,3)).value = total_RC
    ws_set.range((row + 3,4)).value = (total_RC / total_set)
    ws_set.range((row + 4,1)).value = 'OB'
    ws_set.range((row + 4,2)).value = 'Outside Broadcast'
    ws_set.range((row + 4,3)).value = total_OB
    ws_set.range((row + 4,4)).value = (total_OB / total_set)
    ws_set.range((row,4),(row + 4,4)).api.NumberFormat = "0.00%"'''


def report_cast(self, wb_master, all_eps):
    cast_type = list(self.df_cast_report['Type'])
    cast_arr = list(self.df_cast_report['Cast'])
    # it is a single heading for cast report for the first time
    # but for any redo for reporting, it is double heading, so cut first null entry
    if self.df_cast_report['Cast'].isnull().values.any():
        cast_arr = cast_arr[1:len(cast_arr)]
        cast_type = cast_type[1:len(cast_type)]
    # set up report data frame
    team_list = []
    for team in all_eps.keys():
        team_list.append('Team ' + team)
    # prepare df_report for set
    # Create multi header
    header = pd.MultiIndex.from_product([team_list +['Total'], ['#Sc', '#Set']], names=['Team','Count'])
    df_report = pd.DataFrame(columns = header)
    # Add the cast column with cast name
    df_report.insert(0, 'Type', cast_type, True)
    df_report.insert(1, 'Cast', cast_arr, True)
    
    # check data in team master
    for index,team in enumerate(team_list):
        df_team = pd.read_excel(self.master_file, sheet_name= team)
        for row, cast_name in enumerate(cast_arr):
            col = index * 2 + 2
            df_report.iat[row, col] = (df_team.loc[df_team[cast_name] == 'X', ['Set']]).count().to_list()[0]
            col = col + 1
            df_report.iat[row, col] = (df_team.loc[df_team[cast_name] == 'X', ['Set']]).nunique().to_list()[0]
            col = col + 1

    # Calculate total
    for row, cast_name in enumerate(cast_arr):
        total_sc = 0
        total_set = 0
        col = 2
        for index, team in enumerate(team_list):
            total_sc = total_sc + df_report.iat[row, col]
            col = col + 1
            total_set = total_set + df_report.iat[row, col]
            col = col + 1
        df_report.iat[row, col] = total_sc
        col = col + 1
        df_report.iat[row, col] = total_set

    sheets = wb_master.sheetnames
    if 'Cast Report' in sheets:
        del wb_master['Cast Report']
    ws_cast_report = wb_master.create_sheet('Cast Report')
    for r in dataframe_to_rows(df_report, index=False, header=True):
        ws_cast_report.append(r)
    # Format then
    ws_cast_report = format.cast_final(ws_cast_report, team_list)
    # Save
    wb_master.save(self.master_file)

    '''# Format sheet
    ws_cast = wb_master.sheets['Cast Report']
    cast_maxr = ws_cast.range('A3').end('down').row
    cast_maxc = ws_cast.range('A1').end('right').column
    # suppress suppress prompts and alert messages
    wb_master.app.display_alerts = False
    # Set the column width
    ws_cast.range((1, 1),(cast_maxr, 1)).autofit()
    ws_cast.range((3,2),(cast_maxr,cast_maxc)).column_width = 5
    # Set alignment
    ws_cast.range((3,2),(cast_maxr,cast_maxc)).api.HorizontalAlignment = -4108
    # Bold the total values
    ws_cast.range((3,cast_maxc - 1),(cast_maxr,cast_maxc)).api.Font.Bold = True
    # Greying the #Set columns
    col = 3
    for team in team_list:
        ws_cast.range((3,col),(cast_maxr,col)).color = (230, 230, 230)
        ws_cast.range((1,col - 1),(1,col)).merge()
        col = col + 2
    ws_cast.range((3,col),(cast_maxr,col)).color = (230, 230, 230)
    ws_cast.range((1,col - 1),(1,col)).merge()
    row = 3
    # Draw border
    while row < cast_maxr:
        ws_cast.range((row,1),(row,cast_maxc)).api.Borders(8).LineStyle = 1
        ws_cast.range((row,1),(row,cast_maxc)).api.Borders(8).Weight = 2
        row = row + 5
    # disable suppress suppress prompts and alert messages
    wb_master.app.display_alerts = True'''

