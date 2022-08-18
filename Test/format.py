import pandas as pd
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Protection, Font

local_path = ("C:/Users/chanb/OneDrive/Desktop/Scenes to Master V2/test/")
Excelfile = local_path + '101.xlsx'
# Read Excel file to dataframe
df = pd.read_excel(Excelfile)
'''# Create a workbook, and a sheet will auto generated
wb = Workbook()
ws = wb.active'''
# Open master and create eps ws
time_start = time.time()
wb_m = load_workbook(local_path + '101 to 110 master.xlsx')
ws = wb_m.create_sheet('111', -4)
time_master = time.time()
print('Loading master took ' + str(time_master - time_start) + 's')
# Read the df into ws
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
time2pd = time.time()
print('Convert to pd took ' + str(time2pd - time_master) + 's')
# Formatting
CastUsed = 27
lastrow = ws.max_row
lastcolumn = ws.max_column

# Border
for col in ws.iter_cols(min_row=1, max_col= lastcolumn, max_row=lastrow):
    for cell in col:
         cell.border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Row height
ws.row_dimensions[1].height = 88
for row in range(2, lastrow + 1):
    ws.row_dimensions[row].height = 60

col = 1     #Time
ws.column_dimensions[get_column_letter(col)].width = 14.09
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='center', 
            wrap_text=True)
        
col = 2     #Ep
ws.column_dimensions[get_column_letter(col)].width = 4.91
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
        
col = 3     #Sc
ws.column_dimensions[get_column_letter(col)].width = 4.64
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
        
col = 4     #Set
ws.column_dimensions[get_column_letter(col)].width = 28.45
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='left', vertical='top', 
            wrap_text=True)
        
col = 5     #Area
ws.column_dimensions[get_column_letter(col)].width = 31.55
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='left', vertical='top', 
            wrap_text=True)
        
col = 6     #D/N
ws.column_dimensions[get_column_letter(col)].width = 17.09
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
        
col = 7     #Type
ws.column_dimensions[get_column_letter(col)].width = 7.82
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
        
col = 8     #Shoot Time
ws.column_dimensions[get_column_letter(col)].width = 8.18
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.number_format = 'h:mm'
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
        
col = 9     #Pages
ws.column_dimensions[get_column_letter(col)].width = 7.73
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
        
col = 10     #Synopsis
ws.column_dimensions[get_column_letter(col)].width = 47
ws.cell(row = 1, column = col).font = Font(size=14)
ws.cell(row = 1, column = col).alignment=Alignment(horizontal='center',
    vertical='top', wrap_text=True)
for col in ws.iter_cols(min_col= col, max_col= col, min_row=2, max_row=lastrow):
    for cell in col:
        cell.font = Font(size=11)
        cell.alignment=Alignment(horizontal='left', vertical='top', 
            wrap_text=True)

# Now for the cast columns
for col in range(11, 72 + 1):
    ws.column_dimensions[get_column_letter(col)].width = 3.09
for row in ws.iter_rows(min_col=11, max_col=72, min_row=1, max_row= 1):
    for cell in row:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            textRotation=180, wrap_text=True)
for row in ws.iter_rows(min_col=11, max_col=72, min_row=2, max_row=lastrow):
    for cell in row:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)

# Last 3 columns
for col in range(73, 75 + 1):
    ws.column_dimensions[get_column_letter(col)].width = 26.64
for row in ws.iter_rows(min_col=73, max_col=75, min_row=1, max_row= 1):
    for cell in row:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='center', vertical='top', 
            wrap_text=True)
for row in ws.iter_rows(min_col=73, max_col=75, min_row=2, max_row=lastrow):
    for cell in row:
        cell.font = Font(size=14)
        cell.alignment=Alignment(horizontal='left', vertical='top', 
            wrap_text=True)

# Hide unused part timers column
for col in range(11 + CastUsed, 73):
    ws.column_dimensions[get_column_letter(col)].hidden= True
time_format = time.time()
print('Formatting took ' + str(time_format - time2pd) + 's')
# Export the wb to new Excel file
wb_m.save(local_path + '101 to 110 master.xlsx')
time_save = time.time()
print('Saving file took ' + str(time_save - time_format) + 's')
print('Total took ' + str(time_save - time2pd) + 's')